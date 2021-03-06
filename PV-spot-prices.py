import pandas as pd
from datetime import datetime as dt
from pathlib import Path
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import requests
from bs4 import BeautifulSoup as bs

def get_data(url):
    """
    downloads spot price data and its date of update
    """
    response = requests.get(url)
    if response.status_code == 200 :
        soup=bs(response.text,"html.parser")
        date_tag=soup.find_all('span')
        if len(date_tag)>0:
            pass
        else:
            print("Something wrong with getting upload date")
        tables_list=pd.read_html(url)
        if len(tables_list)>0:
            print("Data downloaded successfully")
        else:
            print("Something wrong with downloading Data. Please try again.")
    else:
        print("Request was not successful.")
    return date_tag, tables_list

def check_data_2(wb, date_str):
    """
    based on the update date value checks if the data is relevant.

    """
    for sheet_name in wb.sheetnames:
        if sheet_name == date_str:
            data_status=False
            print("Data is not relevant and is already contained in", filename,"." " File won't be overwritten.")
            break
        else:
            data_status = True
    if data_status==True:
        print("Data is relevant")
    return(data_status)

def datetag_to_datetime(date_tag):
    """
    converts downloaded date tag to a datetime object
    """
    a=str(date_tag).find('>', 0)+1
    b=str(date_tag).find('<', 1)
    string=str(date_tag)[a:b]
    string_cut=string[0:string.find(",")]+string[string.find(",")+1:]
    date=dt.strptime(string_cut, '%B %d %Y')
    return date

def swap_headers(ws, header_old, header_new):
    """
    changes columns names within given worksheet
    """
    for row in ws.rows:
        for cell in row:
            if cell.value==header_old:
                cell.value=header_new

def add_new_data_to_column(ws, df, date_str):
    """
    adds new data to the table
    """
    max_col=ws.max_column
    r=0
    for j in range(len(df.columns)):
        ws.cell(row=4+r, column=max_col+1).value=date_str
        for i in range(len(df.iloc[:,j])):
            ws.cell(row=i+5+r, column=max_col+1).value=df.iloc[:,j][i]
        r=r+i+4

def create_main_tables_structure(ws, df, date_str):
    """
    generates main table structure
    """
    ws['A1']="Data is provided for reference purposes only. Data is the property of PV InfoLink.\nUsers should respect the legitimate rights for use based on the principle of integrity."
    ws['A1'].alignment=Alignment(wrap_text=True)
    ws.merge_cells('A1:D1')
    ws['A2']="Please do not modify this sheet manually. Use a copy for playing with the data"
    ws.merge_cells('A2:D2')
    ws['A3']=df.columns[-3]+", "+ws.title
    ws['A3'].font=Font(bold=True)
    ws.merge_cells('A3:D3')
    for r in dataframe_to_rows(df[["Category", "Item", "Unit", "High Price"]], index=False, header=True):
        ws.append(r)
    ws.cell(row=ws.max_row+2, column=1, value=df.columns[-1]+", "+ws.title)
    ws.cell(row=ws.max_row, column=1).font=Font(bold=True)
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column)
    for r in dataframe_to_rows(df[["Category", "Item", "Unit", "Average Price"]], index=False, header=True):
        ws.append(r)
    ws.cell(row=ws.max_row+2, column=1, value=df.columns[-2]+", "+ws.title)
    ws.cell(row=ws.max_row, column=1).font=Font(bold=True)
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column)
    for r in dataframe_to_rows(df[["Category", "Item", "Unit", "Low Price"]], index=False, header=True):
        ws.append(r)

    ws.cell(row=ws.max_row+2, column=1).value='*The quote of mono wafers is low resistivity product.'
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column)
    ws.cell(row=ws.max_row+1, column=1).value='**Mono-Si wafer quotes are based on those of 180µm. Prices of thinner ones are calculated with formula.'
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column)
    ws.cell(row=ws.max_row+1, column=1).value='***US and Indian module prices showed on the PV InfoLink website is after-tax price (punitive tariffs). Others are FOB price.'
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=ws.max_column)

    swap_headers(ws, "Low Price", date_str)
    swap_headers(ws, "Average Price", date_str)
    swap_headers(ws, "High Price", date_str)
    ws.column_dimensions['B'].width=51
    ws.column_dimensions['A'].width=16
    ws.column_dimensions['D'].width=10
    ws.row_dimensions[1].height=30


#creating output file if needed
filename="Spot_prices.xlsx"
output_file=Path.cwd() / filename
if output_file.is_file():
    wb=load_workbook(filename)
    ws_USD=wb["USD"]
    ws_RMB=wb["RMB"]
    print("Output file", filename, "already exists...\nDownloading Data...")
else:
    wb=Workbook()
    ws_USD=wb.active
    ws_USD.title="USD"
    ws_RMB=wb.create_sheet("RMB",1)
    wb.save(filename)
    print("Output file", filename, "has been created...\nDownloading Data...")

#downloading data
url="https://www.infolink-group.com/en/solar/spot-price"
requested_data=get_data(url)

#getting update date and formating as a datetime object
date_tag=requested_data[0][3]
date=datetag_to_datetime(date_tag)
date_str= date.strftime('%d-%m-%Y')
# dummy_date="11-11-1111"
# date_str=dummy_date

list_of_df=requested_data[1] #getting tables with spot price data

#checking if the data is relevant
data_status = check_data_2(wb, date_str)

#data processing and consolidation
df_all=pd.DataFrame()

for df in list_of_df:
    df_all=df_all.append(df, ignore_index=True)

df_all.drop(["Change(%)", "Change($)", "Price prediction for next week"],axis=1, inplace=True)
df_all.reset_index(drop=True, inplace=True)
df_all.rename(columns={'High':'High Price', 'Low':'Low Price','Average price':'Average Price'},  inplace=True)

category_list=4*["Polysilicon"]+6*["Wafer"]+8*["Cell"]+6*["Module"]+6*["Module by region"]+2*["Module BOM Materials"]
unit_list=4*["kg"]+6*["pc"]+20*["W"]+2*["m2"]

currency_list=[]
for i in range(len(df_all["Item"])):
    currency_list.append(df_all["Item"][i][-4:-1])
# print(currency_list)

df_all.insert(0, "Category", category_list, True)
df_all.insert(2, "Currency", currency_list, True)
df_all.insert(3, "Unit", unit_list, True)


#separating
df_USD=df_all[df_all["Currency"]=="USD"]
df_USD.reset_index(drop=True, inplace=True)
df_RMB=df_all[df_all["Currency"]=="RMB"]
df_RMB.reset_index(drop=True, inplace=True)

#adding new data to a new excel sheet
if data_status == True:
    wb.create_sheet(date_str)
    ws_new=wb[date_str]
    for df in list_of_df:
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_new.append(r)
        ws_new.append([])
    print("Update date:", date_str)
    print("Data is added to the corresponding worksheet. Please check the file.")

#adding new data to main tables (USD, RMB)
if data_status == True:
    if ws_USD['A3'].value==df_USD.columns[-3]+", "+ws_USD.title:
        add_new_data_to_column(ws_USD, df_USD[["High Price","Average Price","Low Price"]],date_str)
        add_new_data_to_column(ws_RMB, df_RMB[["High Price","Average Price","Low Price"]],date_str)
    else:
        create_main_tables_structure(ws_USD, df_USD, date_str)
        create_main_tables_structure(ws_RMB, df_RMB, date_str)

wb.save(filename)

try:
    print("Thanks for using this tool. Ha en hyggelig dag!")
    input("Press enter to exit.")
except SyntaxError:
    pass
